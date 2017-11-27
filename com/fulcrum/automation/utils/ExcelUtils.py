import os
import openpyxl


fileDir = os.path.dirname(os.path.realpath('__file__'))
parentDir = os.path.dirname(fileDir)


class ExcelTestDataAccess:

    def __init__(self, framework_config):
        self.workbook = openpyxl.load_workbook(os.path.join(parentDir,  framework_config.get('testdata.workbook')))

    def get_worksheet_obj(self, sheet_name):
        return self.workbook.get_sheet_by_name(sheet_name)

    @staticmethod
    def get_max_row(worksheet):
        return worksheet.max_row

    @staticmethod
    def get_max_column(worksheet):
        return worksheet.max_column

    def get_row_data(self, sheet_name, current_testcase):

        worksheet = self.get_worksheet_obj(sheet_name)
        total_rows = self.get_max_row(worksheet)
        total_columns = self.get_max_column(worksheet)

        data = {}

        column_names = self.get_column_names(worksheet, total_columns)
        current_row = self.get_row_number(worksheet, total_rows, current_testcase)

        column_counter = 0

        for row in worksheet.iter_cols(min_row=current_row, min_col=1, max_row=current_row, max_col=total_columns):
            for cell in row:
                if cell.value is not None:
                    data[column_names[column_counter]] = cell.value
                    column_counter += 1
        # print data
        return data

    @staticmethod
    def get_row_number(worksheet, total_rows, test_case_name):

        current_row_number = 0

        for row in worksheet.iter_rows(min_row=2, min_col=1, max_row=total_rows, max_col=1):
            for cell in row:
                if cell.value == test_case_name:
                    current_row_number = cell.row

        return current_row_number

    @staticmethod
    def get_column_names(worksheet , total_columns):

        column_names = []

        for row in worksheet.iter_cols(min_row=1, min_col=1, max_row=1, max_col=total_columns):
            for cell in row:
                column_names.append(cell.value)
        return column_names


class ExcelRunManagerAccess:

    def __init__(self, framework_config):
        self.workbook = openpyxl.load_workbook(os.path.join(parentDir, framework_config.get('run.manager')))

    def get_worksheet_obj(self, sheet_name):
        return self.workbook.get_sheet_by_name(sheet_name)

    @staticmethod
    def get_max_row(worksheet):
        return worksheet.max_row

    @staticmethod
    def get_max_column(worksheet):
        return worksheet.max_column

    def get_run_manager_info(self, sheet_name):

        worksheet = self.get_worksheet_obj(sheet_name)
        total_rows = self.get_max_row(worksheet)
        total_columns = self.get_max_column(worksheet)

        run_info = []

        column_names = self.get_column_names(worksheet, total_columns)
        column_counter = 0
        row_counter = 0

        for row in worksheet.iter_rows(min_row=2, min_col=1, max_row=total_rows, max_col=total_columns):
            row_data = {}
            for cell in row:
                if cell.value is not None:
                    row_data[column_names[column_counter]] = cell.value
                    column_counter += 1
            column_counter = 0
            if row_data['EXECUTE'] != 'No':
                run_info.insert(row_counter, row_data)
                row_counter += 1

        # print run_info
        return run_info

    @staticmethod
    def get_column_names(worksheet, total_columns):

        column_names = []

        for row in worksheet.iter_cols(min_row=1, min_col=1, max_row=1, max_col=total_columns):
            for cell in row:
                column_names.append(cell.value)
        return column_names









